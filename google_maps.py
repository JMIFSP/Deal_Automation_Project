from time import sleep

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


def open_driver_connection():
    print("Creating driver connection")
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    print("Driver connection created successfully")
    return driver


def nearby_location(driver, source_link):
    list_item = [
        "Train Stations",
        "Fire Stations",
        "Postal Service",
        "Police Stations",
        "Golf Course",
        "Supermarkets",
        "Distribution Centers",
        "Restaurants",
        "gas stations",
        "Branded convenience stores",
        "Shopping malls",
        "Department stores",
        "Dollar stores",
        "Pharmacy",
        "Outlet malls",
        "Supply Stores",
        "Schools",
        "Universities",
        "Parks",
        "Subway",
        "Shipping and mailing service",
        "Lodge",
        "Hotels",
        "Movie Theaters",
        "Fast food",
        "Beaches",
        "Hospitals",
        "Resorts",
        "Motel",
    ]

    driver.get(source_link)
    WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="pane"]/div/div[1]/div/div/div[7]'))
    )
    address = driver.find_element_by_xpath(
        '//*[@id="pane"]/div/div[1]/div/div/div[7]/div/div[1]/span[3]/span[3]'
    ).text
    print(address)
    driver.find_element_by_xpath(
        '//*[@id="pane"]/div/div[1]/div/div/div[4]/div[3]/div/button'
    ).click()

    google_results = {}
    for name in list_item:
        print("\n========================================================")
        print("name: {}\n".format(name))
        try:
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.ID, "searchboxinput"))
            )
            driver.find_element_by_id("searchboxinput").clear()
            driver.find_element_by_id("searchboxinput").send_keys(name)
            driver.find_element_by_class_name("searchbox-searchbutton-container").enter()
            sleep(5)
            pause_time = 3
            max_count = 6
            x = 0
            while x < max_count:
                if name == "Movie Theaters":
                    x_path = '//*[@id="pane"]/div/div[1]/div/div/div[2]/div[1]'
                    scrollable_div = driver.find_element_by_xpath(x_path)
                elif name in ["Beaches", "Hospitals", "Resorts", "Motel"]:
                    css_selector = (
                        "#pane > div > div.widget-pane-content.cYB2Ge-oHo7ed > div > div > "
                        "div.section-layout.section-scrollbox.cYB2Ge-oHo7ed.cYB2Ge-ti6hGc."
                        "siAUzd-neVct-Q3DXx-vertical > div.section-layout.section-scrollbox."
                        "cYB2Ge-oHo7ed.cYB2Ge-ti6hGc.siAUzd-neVct-Q3DXx-vertical"
                    )
                    scrollable_div = driver.find_element_by_css_selector(css_selector)
                else:
                    x_path = '//*[@id="pane"]/div/div[1]/div/div/div[4]/div[1]'
                    scrollable_div = driver.find_element_by_xpath(x_path)
                try:
                    driver.execute_script(
                        "arguments[0].scrollTop = arguments[0].scrollHeight", scrollable_div
                    )
                except:
                    pass

                sleep(pause_time)
                x = x + 1

            google_links = []
            for elem in driver.find_elements_by_class_name("place-result-container-place-link"):
                google_links.append(elem.get_attribute("href"))

            print("Total: {}".format(len(google_links)))
            if len(google_links) > 0:
                driver.execute_script("window.open('');")
                driver.switch_to.window(driver.window_handles[1])
                current_list = []
                try:
                    for href in google_links:
                        try:
                            # print(href)
                            driver.get(href)

                            WebDriverWait(driver, 80).until(
                                EC.presence_of_element_located(
                                    (By.XPATH, '//*[@id="pane"]/div/div[1]/div/div/div[2]')
                                )
                            )
                            shop_name = driver.find_element_by_xpath(
                                '//*[@id="pane"]/div/div[1]/div/div/div[2]/div[1]/div[1]/div[1]/h1/span[1]'
                            ).text
                            to_address = driver.find_element_by_class_name("AeaXub").text

                            driver.find_element_by_class_name("S9kvJb").click()

                            WebDriverWait(driver, 60).until(
                                EC.presence_of_element_located(
                                    (By.CLASS_NAME, "widget-directions-waypoints")
                                )
                            )
                            driver.find_element_by_xpath('//*[@id="sb_ifc51"]/input').send_keys(
                                address
                            )
                            sleep(2)
                            driver.find_element_by_xpath('//*[@id="sb_ifc51"]/input').send_keys(
                                Keys.ENTER
                            )

                            WebDriverWait(driver, 30).until(
                                EC.presence_of_element_located((By.ID, "section-directions-trip-0"))
                            )
                            try:
                                distance = driver.find_element_by_xpath(
                                    '//*[@id="section-directions-trip-0"]/div/div[1]/div[1]/div[2]/div'
                                ).text
                            except:
                                distance = driver.find_element_by_xpath(
                                    '//*[@id="section-directions-trip-0"]/div/div[3]/div[1]/div[2]'
                                ).text
                            data = {
                                "destination": shop_name,
                                "address": to_address,
                                "distance": distance,
                                "link": href,
                            }
                            current_list.append(data)
                            print(data)
                        except:
                            continue
                finally:
                    google_results[name] = current_list
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
        except Exception as exp:
            print("Failed")
            print(repr(exp))
        finally:
            sleep(2)

    return google_results


def main():
    try:
        driver = open_driver_connection()
        try:
            source_link = (
                "https://www.google.co.in/maps/place/810+Peachtree+St,+Cocoa,+FL+32922,+USA/@28.3577177,-80.7419798,17z/data=!3m1!4b1!4m5!3m4!1s0x88e0aa0a0b8f5687:0xec8d966290351cd6!8m2!3d28.3577177 !4d-80.7397911"
            )
            google_results = nearby_location(driver, source_link)

            file = "/home/user/Documents/google.xlsx"
            files = "/home/user/Documents/google results.xlsx"
            wb = openpyxl.load_workbook(file)
            ws = wb["output"]
            next_row = 2

            for key, values in google_results.items():
                for val in values:
                    ws.cell(column=1, row=next_row, value=key)
                    ws.cell(column=2, row=next_row, value=val["destination"])
                    ws.cell(column=3, row=next_row, value=val["address"])
                    ws.cell(column=4, row=next_row, value=val["distance"])
                    ws.cell(column=5, row=next_row, value=val["link"])
                    next_row += 1
            wb.save(files)
        finally:
            driver.quit()
    except Exception as exp:
        print("Process Failed")
        print(repr(exp))
    else:
        print("Process Ended")


main()
