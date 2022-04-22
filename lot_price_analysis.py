import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


def open_driver_connection():
    print("Creating driver connection")
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    print("Driver connection created successfully")
    return driver


def get_value(driver, url):
    try:
        driver.get(url)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME, "section-layout"))
        )
        driver.execute_script("window.stop();")
        try:
            return driver.find_element_by_xpath(
                '//*[@id="section-directions-trip-0"]/div/div[1]/div[1]/div[2]/div'
            ).text
        except:
            return driver.find_element_by_xpath(
                '//*[@id="section-directions-trip-0"]/div/div[3]/div[1]/div[2]'
            ).text
    except:
        return ""


def main():
    try:
        driver = open_driver_connection()
        try:
            wb = openpyxl.load_workbook(
                "/home/jeanmartin.com/chidambaranathan.m/Desktop/deal_demo/Lot_Price_Analysis.xlsx",
                keep_vba=True,
            )
            for sheet in wb.sheetnames:
                print("\nSheet name: {}\n".format(sheet))
                ws = wb[sheet]
                cell_sheet = ws["O"]
                column_change = ws["M"][0].column
                for col in cell_sheet[1:]:
                    map_link = col.value
                    if map_link:
                        if "#" in map_link:
                            y = map_link.split("#")
                            z = y[1].split("+", 1)
                            map_link = y[0] + z[1]

                        print(map_link)
                        data = get_value(driver, map_link)
                        print(data if data != "" else "No Data found")
                        ws.cell(row=col.row, column=column_change).value = data
            wb.save(
                "/home/jeanmartin.com/chidambaranathan.m/Desktop/deal_demo/Lot_Price_Analysis.xlsx"
            )
        finally:
            driver.quit()
    except Exception as exp:
        print("Process Failed")
        print(repr(exp))
    else:
        print("Process Ended")


main()
