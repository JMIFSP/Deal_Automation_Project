import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from bs4 import BeautifulSoup


def open_driver_connection():
    print("Creating driver connection")
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    print("Driver connection created successfully")
    return driver


def get_vendors_data(driver, zip_code):
    vendor_data = {}
    print("Openning Yelp.com")
    vendors_list = (
        "Electricians",
        "Plumbers",
        "Landscaping",
        "Paving contractors",
        "Fencing contractors",
        "Mobile home moving",
        "Local surveyors",
        "Mobile home dealers",
        "Local real estate brokers",
    )
    for vendor in vendors_list:
        try:
            print("\n====================================")
            print("Vendor: {}\n".format(vendor))
            vendor_data[vendor] = {}

            driver.get("https://www.yelp.com")
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.ID, "header_find_form"))
            )
            driver.execute_script("window.stop();")
            driver.find_element(By.XPATH, "//input[@id='find_desc']").send_keys(vendor)
            driver.find_element(By.XPATH, "//input[@id='dropperText_Mast']").clear()
            driver.find_element(By.XPATH, "//input[@id='dropperText_Mast']").send_keys(zip_code)
            driver.find_element(By.XPATH, "//button[@id='header-search-submit']").click()

            company_links = []
            for elem in driver.find_elements_by_xpath("//*[@class='css-1m051bw']"):
                company_links.append(elem.get_attribute("href"))
            if len(company_links) > 0:
                for href in company_links:
                    try:
                        driver.get(href)
                        WebDriverWait(driver, 60).until(
                            EC.presence_of_element_located((By.CLASS_NAME, 'css-1m051bw'))
                        )
                        driver.execute_script("window.stop();")
                        company_name=driver.find_element(
                            By.XPATH, "//*[@class='css-12dgwvn']"
                        ).text
                        url = driver.current_url
                        print(url)
                        raw_search_page = BeautifulSoup(driver.page_source, features="lxml")
                        phone_number = raw_search_page.find_all("div", {"class": "stickySidebar--fullHeight__09f24__kqHVd arrange-unit__09f24__rqHTg arrange-unit-grid-column--4__09f24__P05hD padding-l2__09f24__kf_t_ border-color--default__09f24__NPAKY"})[0]
                        tag = phone_number.find_all("p")
                        for (index, item) in enumerate(tag):
                            if item.text == "Phone number":
                                x = index + 1
                                no = tag[x].text
                                print(no)
                                vendor_data[vendor].update(
                                    {company_name: {"phone_no": no, "link": url}}
                                )
                                print(vendor_data)
                    except:
                        continue
            else:
                print("Vendor: {} is not available".format(vendor))
        except Exception as exp:
            print("{} data extraction failed".format(vendor))
            print(repr(exp))

    return vendor_data


def main():
    try:
        driver = open_driver_connection()
        try:
            wb = openpyxl.load_workbook(
                "/home/tamilarasi/Backup_280122/Downloads/Vendors_34448 (1).xlsx",
                keep_vba=True,
            )

            # ws = wb["01. Initial Deal Review"]
            # zip_code = ws["C"][10].value
            #zip_code = str(input("Enter the zip_code:"))
            #zip_code = str(sys.argv[0])
            zip_code = 34448
            print("zip_code: {}".format(zip_code))
            vendor_data = get_vendors_data(driver, zip_code)
            print(vendor_data)
            vendor_sheet = wb["DEMO"]
            #vendor_sheet = wb["10. Vendor List"]
            mapping_dict = {
                "Electricians": {"start": 15, "end": 24},
                "Plumbers": {"start": 28, "end": 38},
                "Landscaping": {"start": 42, "end": 50},
                "Paving contractors": {"start": 54, "end": 60},
                "Fencing contractors": {"start": 64, "end": 69},
                "Mobile home moving": {"start": 73, "end": 75},
                "Mobile home dealers": {"start": 80, "end": 83},
                "Local real estate brokers": {"start": 89, "end": 92},
            }
            for map_name, map_value in mapping_dict.items():
                start_key = map_value["start"]
                end_key = map_value["end"]
                for v_name, v_details in vendor_data.get(map_name, {}).items():
                    vendor_sheet.cell(row=start_key, column=2).value = v_name
                    vendor_sheet.cell(row=start_key, column=5).value = v_details["phone_no"]
                    vendor_sheet.cell(row=start_key, column=12).hyperlink = v_details["link"]
                    start_key += 1
                    if start_key > end_key:
                        break
            wb.save(
                "/home/user/Documents/Vendors_34448 (1).xlsx"
            )
        finally:
            driver.quit()
    except Exception as exp:
        print("Process Failed")
        print(repr(exp))
    else:
        print("Process Ended")


if __name__ == "__main__":
    main()