from time import sleep

import js2xml
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


def open_driver_connection():
    print("Creating driver connection")
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    print("Driver connection created successfully")
    return driver


def get_summary(driver, zip_code):
    summary = {}
    try:
        print("demographics summary data extraction is started")
        driver.get("https://www.bestplaces.net/")
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "txtSearch")))
        place_search_box = driver.find_element_by_id("txtSearch")
        place_search_box.clear()
        place_search_box.send_keys(zip_code)
        Go = driver.find_element_by_id("btnSearch")
        Go.click()

        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.CLASS_NAME, "card-body"))
        )
        home_page_data = BeautifulSoup(driver.page_source, "lxml")
        population_tag = home_page_data.find("div", {"class": "card-body container"})
        population_value = population_tag.find("div", {"class": "row"})
        all_cards_data = population_value.find_all("div", {"class": "col-md-4 px-1"})
        first_card = [col_name.text.strip() for col_name in all_cards_data[0].find_all("p")]
        second_card = [col_name.text.strip() for col_name in all_cards_data[1].find_all("p")]
        third_card = [col_name.text.strip() for col_name in all_cards_data[2].find_all("p")]
        # print(first_card)
        # print(second_card)
        # print(third_card)
        population = first_card[1]
        population_growth = first_card[2]
        un_employment = first_card[4]
        median_home_price = second_card[3]
        median_age = third_card[1]
        summary["Population"] = population
        summary["Population growth"] = population_growth
        summary["Unemployment"] = un_employment
        summary["Median Home Price"] = median_home_price
        summary["Median Age"] = median_age
        current_url = driver.current_url
        url_split = current_url.split("zip-code")
        summary["Card Url"] = current_url
        economy_url = url_split[0] + "economy/zip-code" + url_split[1]
        summary["Economy Url"] = economy_url
        housing_url = url_split[0] + "housing/zip-code" + url_split[1]
        summary["Housing Url"] = housing_url

        driver.get(economy_url)
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.ID, "mainContent_dgEconomy"))
        )
        economy_page_data = BeautifulSoup(driver.page_source, "lxml")
        economy_table = economy_page_data.find("table", {"id": "mainContent_dgEconomy"})
        economy_df = pd.read_html(str(economy_table))[0]
        median_household_income = economy_df[economy_df[0] == "Household Income"][1].values[0]
        median_family_income = economy_df[economy_df[0] == "Family Median Income"][1].values[0]
        summary["Median Household Income"] = median_household_income
        summary["Median Family Income"] = median_family_income

        driver.get(housing_url)
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.ID, "mainContent_dgHousing"))
        )
        housing_page_data = BeautifulSoup(driver.page_source, "lxml")
        housing_table = housing_page_data.find("table", {"id": "mainContent_dgHousing"})
        housing_df = pd.read_html(str(housing_table))[0]
        price_increase = housing_df[housing_df[0] == "Home Appr. Last 12 months"][1].values[0]
        vacant_housing = housing_df[housing_df[0] == "Vacant For Rent"][1].values[0]
        summary["1yr Home Price Increase"] = price_increase
        summary["% Rental Availability"] = vacant_housing
        try:
            house_rent = housing_page_data.find_all("div", {"class": "card"})[1].find("script").text
            # print(house_rent)
            rent_card_parsed = js2xml.parse(house_rent)
            # print(rent_card_parsed)
            rent_data = [
                d.xpath(".//array/object/property[@name='y']/number/@value")
                for d in rent_card_parsed.xpath("//property[@name='data']")
            ][0]
            # print(rent_data)
            rent_categories = rent_card_parsed.xpath(
                "//property[@name='categories']//string/text()"
            )
            # print(rent_categories)
            rent_categories = [
                val.replace("-", "").replace(".", "").strip() for val in rent_categories
            ]
            rent_dict = dict(zip(rent_categories, rent_data))
            summary["1BR Rent"] = rent_dict.get("1 Bedroom Home or Apart", "")
            summary["2BR Rent"] = rent_dict.get("2 Bedroom Home or Apart", "")
            summary["3BR Rent"] = rent_dict.get("3 Bedroom Home or Apart", "")

        except Exception as exp:
            print("Housing rent data extraction failed")
            print(repr(exp))
            print("demographics summary data extraction is completed")
    except Exception as exp:
        print("demographics summary data extraction is failed")
        print(repr(exp))

    return summary


def get_city_population_data(driver, postal_code, city):
    city_dict = {}
    try:
        print("{} population data extraction is started".format(city))
        wb_url = "https://worldpopulationreview.com/us-cities/{}-{}-population".format(
            city.replace(" ", "-"), postal_code
        )
        census_url = "http://censusviewer.com/city/{}/{}".format(
            postal_code, city.replace(" ", "+")
        )
        driver.get(wb_url)
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "byPopulation")))
        driver.execute_script("window.stop();")
        current_page = BeautifulSoup(driver.page_source, "lxml")
        city_table = current_page.find("div", {"id": "byPopulation"})
        city_df = pd.read_html(str(city_table))[0]
        city_df.rename(columns={"Annual Growth Rate": "Growth Rate"}, inplace=True)
        city_df = city_df.reindex(columns=["Year", "Population", "Growth Rate"])
        option = [2010, 2018, 2019]
        city_df = city_df[city_df["Year"].isin(option)].reset_index(drop=True)

        # population 2000
        try:
            driver.get(census_url)
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.CLASS_NAME, "data_table"))
            )
            driver.execute_script("window.stop();")
            census_page = BeautifulSoup(driver.page_source, "lxml")
            census_table = census_page.find("table", {"class": "data_table"})
            population_2000 = int(
                census_table.find("tr", {"class": "data_body __total"})
                .find_all("td")[3]
                .text.replace(",", "")
            )
            city_df.loc[len(city_df.index)] = [2000, population_2000, ""]
        except Exception as exp:
            print("{} population data extraction for year 2000 is failed".format(city))
            print(repr(exp))

        city_df = city_df.filter(items=["Year", "Population", "Growth Rate"])
        city_df["Year"] = city_df["Year"].astype(str)
        city_dict = city_df.set_index("Year").to_dict(orient="index")
        city_dict["Wb Url"] = wb_url
        city_dict["Census Url"] = census_url
        print("{} population data extraction is completed".format(city))
    except Exception as exp:
        print("{} population data extraction is failed".format(city))
        print(repr(exp))

    return city_dict


def get_metro_population_data(driver, postal_code, county):
    msa_dict = {}
    try:
        print("{} population data extraction is started".format(county))
        wb_url = "https://worldpopulationreview.com/us-counties/{}/{}-county-population".format(
            postal_code, county.replace(" ", "-")
        )
        census_url = "http://censusviewer.com/county/{}/{}".format(
            postal_code, county.replace(" ", "+")
        )
        driver.get(wb_url)
        msa_current_page = BeautifulSoup(driver.page_source, "lxml")
        msa_table = msa_current_page.find("div", {"class": "jsx-2642336383 table-container"})
        msa_df = pd.read_html(str(msa_table))[0]
        msa_df = msa_df.reindex(columns=["Year", "Population", "Growth Rate"])
        option = [1990, 2010, 2018, 2019]
        msa_df = msa_df[msa_df["Year"].isin(option)].reset_index(drop=True)

        # population 2000
        try:
            driver.get(census_url)
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.CLASS_NAME, "data_table"))
            )
            driver.execute_script("window.stop();")
            census_page = BeautifulSoup(driver.page_source, "lxml")
            census_table = census_page.find("table", {"class": "data_table"})
            population_2000 = int(
                census_table.find("tr", {"class": "data_body __total"})
                .find_all("td")[3]
                .text.replace(",", "")
            )
            msa_df.loc[len(msa_df.index)] = [2000, population_2000, ""]
        except Exception as exp:
            print("{} population data extraction for year 2000 is failed".format(county))
            print(repr(exp))

        msa_df = msa_df.filter(items=["Year", "Population", "Growth Rate"])
        msa_df["Year"] = msa_df["Year"].astype(str)
        msa_dict = msa_df.set_index("Year").to_dict(orient="index")
        msa_dict["Wb Url"] = wb_url
        msa_dict["Census Url"] = census_url
        print("{} population data extraction is completed".format(county))
    except Exception as exp:
        print("{} population data extraction is failed".format(county))
        print(repr(exp))

    return msa_dict


def unemployment(driver, postal_code, city, state, county):
    un_employment = {}
    try:
        print("Unemployment data extraction is started")
        employment_url = "https://www.areavibes.com/{}-{}/employment/".format(
            city.replace(" ", "+"), postal_code
        )
        driver.get(employment_url)
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, "av-default"))
        )
        employment_page = BeautifulSoup(driver.page_source, "lxml")
        employment_table = employment_page.find("table", {"class": "av-default"})
        employment_df = pd.read_html(str(employment_table))[0]
        un_employment["City"] = employment_df[employment_df[0] == "Unemployment rate"][1].values[0]
        un_employment["State"] = employment_df[employment_df[0] == "Unemployment rate"][2].values[0]
        un_employment["National"] = employment_df[employment_df[0] == "Unemployment rate"][
            3
        ].values[0]
        un_employment["Unemployment Data Url"] = employment_url
        # Get MSA data
        try:
            msa_url = "https://fred.stlouisfed.org/categories/27281"
            driver.get(msa_url)
            # select state
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.CLASS_NAME, "states"))
            )
            state_list = driver.find_element_by_xpath(
                "/html/body/div/div[1]/div[2]/div[2]/div/div[2]"
            )
            for option in state_list.find_elements_by_tag_name("li"):
                if state.lower() == option.text.split("(")[0].strip().lower():
                    option.find_element_by_tag_name("a").click()
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "series-pager"))
                    )
                    break

            # click county in state page
            driver.find_element_by_xpath(
                "/html/body/div/div[1]/div[2]/div[2]/div/form/div[2]/div[1]/div[1]/ul/li[1]/a"
            ).click()

            # select county
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.CLASS_NAME, "counties"))
            )
            county_list = driver.find_element_by_xpath(
                "/html/body/div/div[1]/div[2]/div[2]/div/div[2]"
            )
            for option in county_list.find_elements_by_tag_name("li"):
                if county.lower() in option.text.strip().lower():
                    option.find_element_by_tag_name("a").click()
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "series-pager"))
                    )
                    break

            # select and click unemployment rate link in county page
            for tr in driver.find_elements_by_class_name("series-pager-title"):
                if "unemployment rate" in tr.text.lower():
                    tr.find_element_by_tag_name("a").click()
                    WebDriverWait(driver, 60).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "series-page"))
                    )
                    break
            un_employment_msa_url = driver.current_url

            # get unemployment rate
            msa_unemployment = (
                driver.find_element_by_xpath(
                    "/html/body/div[1]/div[1]/div/div[2]/div[1]/div[2]/span[2]"
                ).text
                + "%"
            )

        except Exception as exp:
            print("MSA unemployment data extraction failed")
            print(repr(exp))
            msa_unemployment = ""
            un_employment_msa_url = ""
        un_employment["MSA"] = msa_unemployment
        un_employment["Msa Url"] = un_employment_msa_url
        print("Unemployment data extraction is completed")
    except Exception as exp:
        print("Unemployment data extraction is failed")
        print(repr(exp))

    return un_employment


def get_income_property_and_rents(driver, postal_code, city, state, county):
    income_dict = {}
    mapping_dict = {}
    try:
        city = city.lower()
        state = state.lower()
        county = county.lower()
        print("Income and property data extraction is started")
        url = "https://www.census.gov/quickfacts/fact/table/US,{}/".format(postal_code)
        driver.get(url)
        for name in [city, county]:
            search_box = driver.find_element_by_id("qf-search-box")
            search_box.clear()
            search_box.send_keys(name)
            search_box.click()
            sleep(5)
            for option in driver.find_elements_by_xpath("//li[@class='ui-menu-item']//div"):
                option_text = option.text.lower()
                if name == city:
                    validation_list = [city, state]
                else:
                    validation_list = [county, "county", state]

                valid_count = 0
                for y in validation_list:
                    if y in option_text:
                        valid_count += 1

                if valid_count == len(validation_list):
                    mapping_dict[name] = option.text
                    option.click()
                    sleep(5)
                    break

        raw_search_page = BeautifulSoup(driver.page_source, features="lxml")
        all_tables = raw_search_page.find_all("table", {"class": "type"})[0:2]
        header_columns = [
            col_name.text.strip()
            for col_name in all_tables[0].find_all("div", {"class": "qf-geobox"})
        ]
        median_household_income = [
            int(col_value.text.strip().replace("$", "").replace(",", ""))
            for col_value in all_tables[1]
            .find("tbody", {"data-topic": "Income & Poverty"})
            .find_all("tr")[1]
            .find_all("td")[1:]
        ]
        median_value_sfr_homes = [
            int(col_value.text.strip().replace("$", "").replace(",", ""))
            for col_value in all_tables[1]
            .find("tbody", {"data-topic": "Housing"})
            .find_all("tr")[3]
            .find_all("td")[1:]
        ]
        median_gross_rent = [
            int(col_value.text.strip().replace("$", "").replace(",", ""))
            for col_value in all_tables[1]
            .find("tbody", {"data-topic": "Housing"})
            .find_all("tr")[6]
            .find_all("td")[1:]
        ]
        income_Url = driver.current_url
        income_dict["Median Household Income"] = dict(zip(header_columns, median_household_income))
        income_dict["Median Value of SFR Homes"] = dict(zip(header_columns, median_value_sfr_homes))
        income_dict["Median Gross Rent"] = dict(zip(header_columns, median_gross_rent))
        income_dict["Income & Poverty Url"] = income_Url
        print("Income and property data extraction is completed")
    except Exception as exp:
        print("Income and property data extraction is failed")
        print(repr(exp))

    return income_dict, mapping_dict


def crime_data(driver, postal_code, city):
    try:
        url = "https://www.areavibes.com/{}-{}/crime/".format(city.replace(" ", "+"), postal_code)
        driver.get(url)
        soup = BeautifulSoup(driver.page_source, "lxml")
        cities = soup.find("div", {"class": "section spaced-out"})
        city = cities.find("div", {"class": "twelve columns"})
        citi = city.find("div", {"class": "table-overflow-container"})
        city_count = citi.find("tr", {"class": "summary"}).find_all("td")[3:6]
        city_counts = city_count[0].text.split(" ")
        city_counts1 = city_counts[0]
        state = city_count[1].text
        national = city_count[2].text
        crime = {"Crime Url": url, "City": city_counts1, "State": state, "National": national}
    except Exception as exp:
        print(repr(exp))
        crime = {}

    return crime


def get_postal_code(driver, state):
    driver.get(
        "https://www.infoplease.com/us/postal-information/state-abbreviations-and-state-postal-codes"
    )
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "sgmltable")))
    driver.execute_script("window.stop();")
    postal_page = BeautifulSoup(driver.page_source, "lxml")
    postal_table = postal_page.find("table", {"class": "sgmltable"})
    postal_df = pd.read_html(str(postal_table))[0]
    state = postal_df[postal_df["Postal Code"] == state]["State/District"].values[0]
    return state


def main():
    try:
        wb = openpyxl.load_workbook(
            "/home/tamilarasi/Backup_280122/Downloads/GB - Initial Deal Report & DD Worksheet.xlsx",
            keep_vba=True,
        )
        ws = wb["01. Initial Deal Review"]
        print(ws.title)
        city = ws["C"][8].value
        postal_code = ws["C"][9].value
        zip_Code = ws["C"][10].value
        county = ws["C"][11].value
        if "County" in county:
            county = county.replace("County", "").strip()

        print("city: {}".format(city))
        print("Postal Code: {}".format(postal_code))
        print("zip_code: {}".format(zip_Code))
        print("county: {}".format(county))
        driver = open_driver_connection()
        try:
            state = get_postal_code(driver, postal_code)
            print("State: {}".format(state))
            summary = get_summary(driver, zip_Code)
            print(summary)
            un_employment = unemployment(driver, postal_code, city, state, county)
            print(un_employment)
            city_dict = get_city_population_data(driver, postal_code, city)
            print(city_dict)
            msa_dict = get_metro_population_data(driver, postal_code, county)
            print(msa_dict)
            income_dict, mapping_dict = get_income_property_and_rents(
                driver, postal_code, city, state, county
            )
            print(income_dict)
            print(mapping_dict)
            crime = crime_data(driver, city, postal_code)
            print(crime)
            print("writing to excel")

            cell_sheet = wb["02. Demographics"]
            cell_sheet["D12"] = summary.get("Population", "")
            cell_sheet["D13"] = summary.get("Population growth", "")
            cell_sheet["D14"] = summary.get("Unemployment", "")
            cell_sheet["D15"] = summary.get("Median Age", "")
            cell_sheet["D16"] = summary.get("Median Household Income", "")
            cell_sheet["D17"] = summary.get("Median Family Income", "")
            cell_sheet["H12"] = summary.get("1BR Rent", "")
            cell_sheet["H13"] = summary.get("2BR Rent", "")
            cell_sheet["H14"] = summary.get("3BR Rent", "")
            cell_sheet["H15"] = summary.get("Median Home Price", "")
            cell_sheet["H16"] = summary.get("1yr Home Price Increase", "")
            cell_sheet["H17"] = summary.get("% Rental Availability", "")
            cell_sheet["K13"].hyperlink = summary.get("Card Url", "")
            cell_sheet["K13"].value = "Bestplaces"
            cell_sheet["K13"].style = "Hyperlink"
            cell_sheet["K14"].hyperlink = summary.get("Economy Url", "")
            cell_sheet["K14"].value = "Economy"
            cell_sheet["K14"].style = "Hyperlink"
            cell_sheet["K15"].hyperlink = summary.get("Housing Url", "")
            cell_sheet["K15"].value = "Housing"
            cell_sheet["K15"].style = "Hyperlink"
            cell_sheet["D25"] = city_dict.get("2000", {}).get("Population", "")
            # cell_sheet["E25"] = city_dict.get("2010", {}).get("Growth Rate", "")
            cell_sheet["D26"] = city_dict.get("2010", {}).get("Population", "")
            cell_sheet["D27"] = city_dict.get("2018", {}).get("Population", "")
            # cell_sheet["E27"] = city_dict.get("2018", {}).get("Growth Rate", "")
            cell_sheet["D28"] = city_dict.get("2019", {}).get("Population", "")
            # cell_sheet["E28"] = city_dict.get("2019", {}).get("Growth Rate", "")
            cell_sheet["F25"].hyperlink = city_dict.get("Census Url", "")
            cell_sheet["F25"].value = "Link"
            cell_sheet["F25"].style = "Hyperlink"
            cell_sheet["F26"].hyperlink = city_dict.get("Wb Url", "")
            cell_sheet["F26"].value = "Link"
            cell_sheet["F26"].style = "Hyperlink"
            cell_sheet["F27"].hyperlink = city_dict.get("Wb Url", "")
            cell_sheet["F27"].value = "Link"
            cell_sheet["F27"].style = "Hyperlink"
            cell_sheet["F28"].hyperlink = city_dict.get("Wb Url", "")
            cell_sheet["F28"].value = "Link"
            cell_sheet["F28"].style = "Hyperlink"
            cell_sheet["D31"] = msa_dict.get("1990", {}).get("Population", "")
            cell_sheet["D32"] = msa_dict.get("2000", {}).get("Population", "")
            # cell_sheet["E32"] = msa_dict.get("2000", {}).get("Growth Rate", "")
            cell_sheet["D33"] = msa_dict.get("2010", {}).get("Population", "")
            # cell_sheet["E33"] = msa_dict.get("2010", {}).get("Growth Rate", "")
            cell_sheet["D34"] = msa_dict.get("2018", {}).get("Population", "")
            # cell_sheet["E34"] = msa_dict.get("2018", {}).get("Growth Rate", "")
            cell_sheet["D35"] = msa_dict.get("2019", {}).get("Population", "")
            # cell_sheet["E35"] = msa_dict.get("2019", {}).get("Growth Rate", "")
            cell_sheet["F31"].hyperlink = msa_dict.get("Wb Url", "")
            cell_sheet["F31"].value = "Link"
            cell_sheet["F31"].style = "Hyperlink"
            cell_sheet["F32"].hyperlink = msa_dict.get("Census Url", "")
            cell_sheet["F32"].value = "Link"
            cell_sheet["F32"].style = "Hyperlink"
            cell_sheet["F33"].hyperlink = msa_dict.get("Wb Url", "")
            cell_sheet["F33"].value = "Link"
            cell_sheet["F33"].style = "Hyperlink"
            cell_sheet["F34"].hyperlink = msa_dict.get("Wb Url", "")
            cell_sheet["F34"].value = "Link"
            cell_sheet["F34"].style = "Hyperlink"
            cell_sheet["F35"].hyperlink = msa_dict.get("Wb Url", "")
            cell_sheet["F35"].value = "Link"
            cell_sheet["F35"].style = "Hyperlink"
            cell_sheet["D40"] = un_employment.get("City", "")
            cell_sheet["D41"] = un_employment.get("MSA", "")
            cell_sheet["D42"] = un_employment.get("State", "")
            cell_sheet["D43"] = un_employment.get("National", "")
            cell_sheet["F40"].hyperlink = un_employment.get("Unemployment Data Url", "")
            cell_sheet["F40"].value = "Link"
            cell_sheet["F40"].style = "Hyperlink"
            cell_sheet["F41"].hyperlink = un_employment.get("Msa Url", "")
            cell_sheet["F41"].value = "Link"
            cell_sheet["F41"].style = "Hyperlink"
            cell_sheet["F42"].hyperlink = un_employment.get("Unemployment Data Url", "")
            cell_sheet["F42"].value = "Link"
            cell_sheet["F42"].style = "Hyperlink"
            cell_sheet["F43"].hyperlink = un_employment.get("Unemployment Data Url", "")
            cell_sheet["F43"].value = "Link"
            cell_sheet["F43"].style = "Hyperlink"

            cell_sheet["J24"] = income_dict.get("Median Household Income", {}).get(
                mapping_dict.get(county.lower(), "")
            )
            cell_sheet["K24"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K24"].value = "link"
            cell_sheet["K24"].style = "Hyperlink"
            cell_sheet["J25"] = income_dict.get("Median Household Income", {}).get(
                mapping_dict.get(city.lower(), "")
            )
            cell_sheet["K25"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K25"].value = "link"
            cell_sheet["K25"].style = "Hyperlink"
            cell_sheet["J26"] = income_dict.get("Median Household Income", {}).get(state, "")
            cell_sheet["K26"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K26"].value = "link"
            cell_sheet["K26"].style = "Hyperlink"
            cell_sheet["J27"] = income_dict.get("Median Household Income", {}).get(
                "United States", ""
            )
            cell_sheet["K27"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K27"].value = "link"
            cell_sheet["K27"].style = "Hyperlink"
            cell_sheet["J30"] = income_dict.get("Median Value of SFR Homes", {}).get(
                mapping_dict.get(county.lower(), "")
            )
            cell_sheet["K30"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K30"].value = "link"
            cell_sheet["K30"].style = "Hyperlink"
            cell_sheet["J31"] = income_dict.get("Median Value of SFR Homes", {}).get(
                mapping_dict.get(city.lower(), "")
            )
            cell_sheet["K31"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K31"].value = "link"
            cell_sheet["K31"].style = "Hyperlink"
            cell_sheet["J32"] = income_dict.get("Median Value of SFR Homes", {}).get(state, "")
            cell_sheet["K32"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K32"].value = "link"
            cell_sheet["K32"].style = "Hyperlink"
            cell_sheet["J33"] = income_dict.get("Median Value of SFR Homes", {}).get(
                "United States", ""
            )
            cell_sheet["K33"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K33"].value = "link"
            cell_sheet["K33"].style = "Hyperlink"
            cell_sheet["J36"] = income_dict.get("Median Gross Rent", {}).get(
                mapping_dict.get(county.lower(), "")
            )
            cell_sheet["K36"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K36"].value = "link"
            cell_sheet["K36"].style = "Hyperlink"
            cell_sheet["J37"] = income_dict.get("Median Gross Rent", {}).get(
                mapping_dict.get(city.lower(), "")
            )
            cell_sheet["K37"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K37"].value = "link"
            cell_sheet["K37"].style = "Hyperlink"
            cell_sheet["J38"] = income_dict.get("Median Gross Rent", {}).get(state, "")
            cell_sheet["K38"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K38"].value = "link"
            cell_sheet["K38"].style = "Hyperlink"
            cell_sheet["J39"] = income_dict.get("Median Gross Rent", {}).get("United States", "")
            cell_sheet["K39"].hyperlink = income_dict.get("Income & Poverty Url", "")
            cell_sheet["K39"].value = "link"
            cell_sheet["K39"].style = "Hyperlink"
            # cell_sheet["G44:H44"].hyperlink = crime.get("Crime Url", "")
            # cell_sheet["G44:H44"].value = "Areavibes"
            # cell_sheet["G44:H44"].style = "Hyperlink"
            cell_sheet["I44"] = crime.get("City", "")
            cell_sheet["J44"] = crime.get("State", "")
            cell_sheet["K44"] = crime.get("National", "")
            wb.save(
                "/home/tamilarasi/Backup_280122/Downloads/GB - Initial Deal Report & DD Worksheet.xlsx"
            )
        finally:
            driver.quit()
    except Exception as exp:
        print("Process Failed")
        print(repr(exp))
    else:
        print("Process Ended")


if __name__ == "__main__":
    main()
